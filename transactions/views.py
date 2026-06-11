from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.generic import CreateView, UpdateView, DeleteView, ListView
from django.urls import reverse_lazy
from django.db.models import Sum, Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
import json

from .models import Transaction, Category
from .forms import (
    ExpenseForm, IncomeForm, TransferForm,
    CategoryForm, TransactionFilterForm
)
from .services import get_recent_transactions
from accounts.models import BankAccount


# ─── EXPENSE VIEWS ───────────────────────────────────────────────

def expense_view(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            if request.headers.get('HX-Request'):
                recent = Transaction.objects.filter(
                    transaction_type='expense'
                ).select_related('bank_account', 'category').order_by(
                    '-transaction_date', '-created_at'
                )[:10]
                return render(request, 'partials/expense_success.html', {
                    'transaction': transaction,
                    'recent_transactions': recent,
                    'form': ExpenseForm(),
                })
            messages.success(request, f'Expense of ₹{transaction.amount} recorded!')
            return redirect('transactions:expense')
        else:
            if request.headers.get('HX-Request'):
                return render(request, 'partials/form_errors.html', {'form': form})
    else:
        form = ExpenseForm()

    recent = Transaction.objects.filter(
        transaction_type='expense'
    ).select_related('bank_account', 'category').order_by(
        '-transaction_date', '-created_at'
    )[:10]

    context = {
        'form': form,
        'recent_transactions': recent,
        'page_title': 'Add Expense',
    }
    return render(request, 'transactions/expense.html', context)


# ─── INCOME VIEWS ────────────────────────────────────────────────

def income_view(request):
    if request.method == 'POST':
        form = IncomeForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            if request.headers.get('HX-Request'):
                recent = Transaction.objects.filter(
                    transaction_type='income'
                ).select_related('bank_account', 'category').order_by(
                    '-transaction_date', '-created_at'
                )[:10]
                return render(request, 'partials/income_success.html', {
                    'transaction': transaction,
                    'recent_transactions': recent,
                    'form': IncomeForm(),
                })
            messages.success(request, f'Income of ₹{transaction.amount} recorded!')
            return redirect('transactions:income')
        else:
            if request.headers.get('HX-Request'):
                return render(request, 'partials/form_errors.html', {'form': form})
    else:
        form = IncomeForm()

    recent = Transaction.objects.filter(
        transaction_type='income'
    ).select_related('bank_account', 'category').order_by(
        '-transaction_date', '-created_at'
    )[:10]

    context = {
        'form': form,
        'recent_transactions': recent,
        'page_title': 'Add Income',
    }
    return render(request, 'transactions/income.html', context)


# ─── TRANSFER VIEWS ──────────────────────────────────────────────

def transfer_view(request):
    if request.method == 'POST':
        form = TransferForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            if request.headers.get('HX-Request'):
                recent = Transaction.objects.filter(
                    transaction_type='transfer'
                ).select_related('bank_account', 'target_account').order_by(
                    '-transaction_date', '-created_at'
                )[:10]
                return render(request, 'partials/transfer_success.html', {
                    'transaction': transaction,
                    'recent_transactions': recent,
                    'form': TransferForm(),
                    'accounts': BankAccount.objects.filter(is_active=True),
                })
            messages.success(
                request,
                f'Transfer of ₹{transaction.amount} from '
                f'{transaction.bank_account.name} to '
                f'{transaction.target_account.name} successful!'
            )
            return redirect('transactions:transfer')
        else:
            if request.headers.get('HX-Request'):
                return render(request, 'partials/form_errors.html', {'form': form})
    else:
        form = TransferForm()

    recent = Transaction.objects.filter(
        transaction_type='transfer'
    ).select_related('bank_account', 'target_account').order_by(
        '-transaction_date', '-created_at'
    )[:10]

    accounts = BankAccount.objects.filter(is_active=True)
    context = {
        'form': form,
        'recent_transactions': recent,
        'accounts': accounts,
        'page_title': 'Transfer Funds',
    }
    return render(request, 'transactions/transfer.html', context)


# ─── TRANSACTION LIST ────────────────────────────────────────────

def transaction_list(request):
    form = TransactionFilterForm(request.GET)
    queryset = Transaction.objects.select_related(
        'bank_account', 'target_account', 'category'
    ).order_by('-transaction_date', '-created_at')

    if form.is_valid():
        data = form.cleaned_data
        if data.get('search'):
            queryset = queryset.filter(
                Q(note__icontains=data['search']) |
                Q(category__name__icontains=data['search']) |
                Q(bank_account__name__icontains=data['search'])
            )
        if data.get('transaction_type'):
            queryset = queryset.filter(transaction_type=data['transaction_type'])
        if data.get('category'):
            queryset = queryset.filter(category=data['category'])
        if data.get('account'):
            queryset = queryset.filter(
                Q(bank_account=data['account']) | Q(target_account=data['account'])
            )
        if data.get('date_from'):
            queryset = queryset.filter(transaction_date__gte=data['date_from'])
        if data.get('date_to'):
            queryset = queryset.filter(transaction_date__lte=data['date_to'])

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(queryset, 25)
    page = request.GET.get('page', 1)
    transactions = paginator.get_page(page)

    total_income = queryset.filter(transaction_type='income').aggregate(
        t=Sum('amount'))['t'] or 0
    total_expense = queryset.filter(transaction_type='expense').aggregate(
        t=Sum('amount'))['t'] or 0

    if request.headers.get('HX-Request'):
        return render(request, 'partials/transaction_table.html', {
            'transactions': transactions,
            'paginator': paginator,
        })

    context = {
        'form': form,
        'transactions': transactions,
        'total_income': total_income,
        'total_expense': total_expense,
        'page_title': 'All Transactions',
    }
    return render(request, 'transactions/transaction_list.html', context)


# ─── TRANSACTION EDIT / DELETE ───────────────────────────────────

def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    if transaction.transaction_type == 'expense':
        FormClass = ExpenseForm
        template = 'transactions/expense_edit.html'
    elif transaction.transaction_type == 'income':
        FormClass = IncomeForm
        template = 'transactions/income_edit.html'
    else:
        FormClass = TransferForm
        template = 'transactions/transfer_edit.html'

    if request.method == 'POST':
        form = FormClass(request.POST, instance=transaction)
        if form.is_valid():
            t = form.save()
            messages.success(request, 'Transaction updated successfully!')
            return redirect('transactions:list')
    else:
        form = FormClass(instance=transaction)

    return render(request, template, {'form': form, 'transaction': transaction})


@require_POST
def transaction_delete(request, pk):
    from django.utils.http import url_has_allowed_host_and_scheme
    transaction = get_object_or_404(Transaction, pk=pk)
    amount = transaction.amount
    t_type = transaction.transaction_type
    transaction.delete()

    if request.headers.get('HX-Request'):
        return HttpResponse(
            f'<div class="alert alert-success alert-dismissible fade show" role="alert">'
            f'Transaction of ₹{amount} ({t_type}) deleted successfully!'
            f'<button type="button" class="btn-close" data-bs-dismiss="alert"></button>'
            f'</div>'
        )
    messages.success(request, f'Transaction of ₹{amount} ({t_type}) deleted successfully!')
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('transactions:list')


# ─── CATEGORY VIEWS ──────────────────────────────────────────────

def category_list(request):
    income_categories = Category.objects.filter(category_type='income')
    expense_categories = Category.objects.filter(category_type='expense')
    form = CategoryForm()

    context = {
        'income_categories': income_categories,
        'expense_categories': expense_categories,
        'form': form,
        'page_title': 'Categories',
    }
    return render(request, 'transactions/category_list.html', context)


def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            if request.headers.get('HX-Request'):
                income_categories = Category.objects.filter(category_type='income')
                expense_categories = Category.objects.filter(category_type='expense')
                return render(request, 'partials/category_list_partial.html', {
                    'income_categories': income_categories,
                    'expense_categories': expense_categories,
                    'message': f'Category "{category.name}" created!',
                    'form': CategoryForm(),
                })
            messages.success(request, f'Category "{category.name}" created!')
            return redirect('transactions:categories')
        else:
            if request.headers.get('HX-Request'):
                return render(request, 'partials/form_errors.html', {'form': form})
            # Surface validation errors instead of silently redirecting
            error_text = '; '.join(
                f'{form.fields[field].label or field}: {", ".join(errs)}'
                if field in form.fields else ", ".join(errs)
                for field, errs in form.errors.items()
            )
            messages.error(request, f'Could not add category. {error_text}')
    return redirect('transactions:categories')


def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated!')
            return redirect('transactions:categories')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'transactions/category_form.html', {
        'form': form, 'category': category
    })


@require_POST
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    name = category.name
    category.delete()
    messages.success(request, f'Category "{name}" deleted!')
    return redirect('transactions:categories')


# ─── IMPORT / EXPORT ─────────────────────────────────────────────

def export_excel(request):
    """Export all transactions to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Transactions Sheet ──
    ws = wb.active
    ws.title = 'Transactions'

    headers = [
        'Date', 'Type', 'Category', 'Account',
        'Target Account', 'Amount', 'Note', 'Created At'
    ]

    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    transactions = Transaction.objects.select_related(
        'bank_account', 'target_account', 'category'
    ).order_by('-transaction_date')

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=str(t.transaction_date))
        ws.cell(row=row, column=2, value=t.transaction_type.capitalize())
        ws.cell(row=row, column=3, value=t.category.name if t.category else '')
        ws.cell(row=row, column=4, value=t.bank_account.name)
        ws.cell(row=row, column=5, value=t.target_account.name if t.target_account else '')
        ws.cell(row=row, column=6, value=float(t.amount))
        ws.cell(row=row, column=7, value=t.note)
        ws.cell(row=row, column=8, value=str(t.created_at.strftime('%Y-%m-%d %H:%M')))

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Accounts Sheet ──
    ws2 = wb.create_sheet('Accounts')
    acc_headers = ['Name', 'Type', 'Balance', 'Color', 'Icon', 'Active']
    for col, h in enumerate(acc_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, acc in enumerate(BankAccount.objects.all(), 2):
        ws2.cell(row=row, column=1, value=acc.name)
        ws2.cell(row=row, column=2, value=acc.account_type)
        ws2.cell(row=row, column=3, value=float(acc.balance))
        ws2.cell(row=row, column=4, value=acc.color)
        ws2.cell(row=row, column=5, value=acc.icon)
        ws2.cell(row=row, column=6, value='Yes' if acc.is_active else 'No')

    # ── Response ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="finance_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


def export_csv(request):
    """Export transactions to CSV"""
    import csv

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d")}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Type', 'Category', 'Account',
        'Target Account', 'Amount', 'Note'
    ])

    transactions = Transaction.objects.select_related(
        'bank_account', 'target_account', 'category'
    ).order_by('-transaction_date')

    for t in transactions:
        writer.writerow([
            str(t.transaction_date),
            t.transaction_type,
            t.category.name if t.category else '',
            t.bank_account.name,
            t.target_account.name if t.target_account else '',
            str(t.amount),
            t.note,
        ])

    return response


def import_excel(request):
    """Import transactions from Excel"""
    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next')
    if not (next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()})):
        next_url = 'transactions:import_export'
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file.')
            return redirect(next_url)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            imported = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue
                try:
                    date_val = row[0]
                    t_type = str(row[1]).lower() if row[1] else 'expense'
                    category_name = str(row[2]) if row[2] else None
                    account_name = str(row[3]) if row[3] else None
                    target_name = str(row[4]) if row[4] else None
                    amount = float(row[5]) if row[5] else 0
                    note = str(row[6]) if row[6] else ''

                    if not account_name or amount <= 0:
                        skipped += 1
                        continue

                    account = BankAccount.objects.filter(name=account_name).first()
                    if not account:
                        account = BankAccount.objects.create(
                            name=account_name,
                            account_type='savings',
                            balance=0
                        )

                    target_account = None
                    if target_name and t_type == 'transfer':
                        target_account = BankAccount.objects.filter(name=target_name).first()

                    category = None
                    if category_name and t_type in ('income', 'expense'):
                        cat_type = 'income' if t_type == 'income' else 'expense'
                        category = Category.objects.filter(
                            name=category_name, category_type=cat_type
                        ).first()
                        if not category:
                            category = Category.objects.create(
                                name=category_name, category_type=cat_type
                            )

                    from datetime import date as date_type
                    if isinstance(date_val, str):
                        from dateutil.parser import parse
                        t_date = parse(date_val).date()
                    elif isinstance(date_val, date_type):
                        t_date = date_val
                    else:
                        t_date = timezone.now().date()

                    Transaction.objects.create(
                        transaction_type=t_type,
                        bank_account=account,
                        target_account=target_account,
                        category=category,
                        amount=amount,
                        note=note,
                        transaction_date=t_date,
                    )
                    imported += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            msg = f'Import complete! {imported} imported, {skipped} skipped.'
            if errors:
                msg += f' {len(errors)} errors.'
            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')

        return redirect(next_url)

    return redirect(next_url)


def import_export_view(request):
    """Import/Export page"""
    total_transactions = Transaction.objects.count()
    total_accounts = BankAccount.objects.count()
    context = {
        'total_transactions': total_transactions,
        'total_accounts': total_accounts,
        'page_title': 'Import / Export',
    }
    return render(request, 'transactions/import_export.html', context)


def backup_database(request):
    """Download SQLite database backup"""
    import shutil
    from django.conf import settings
    import os

    db_path = settings.DATABASES['default']['NAME']
    backup_name = f'finance_backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.sqlite3'

    with open(db_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{backup_name}"'
    return response


def restore_database(request):
    """Restore database from backup"""
    if request.method == 'POST':
        backup_file = request.FILES.get('backup_file')
        if not backup_file:
            messages.error(request, 'Please select a backup file.')
            return redirect('transactions:import_export')

        try:
            from django.conf import settings
            db_path = settings.DATABASES['default']['NAME']

            # Write uploaded file to db location
            with open(db_path, 'wb') as f:
                for chunk in backup_file.chunks():
                    f.write(chunk)

            messages.success(request, 'Database restored successfully! Please refresh.')
        except Exception as e:
            messages.error(request, f'Restore failed: {str(e)}')

    return redirect('transactions:import_export')