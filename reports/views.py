from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg
from django.http import HttpResponse
import calendar
import json
from decimal import Decimal

from transactions.models import Transaction, Category
from accounts.models import BankAccount


def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError


def reports_view(request):
    now = timezone.now()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))
    report_type = request.GET.get('report_type', 'monthly')

    context = {
        'year': year,
        'month': month,
        'report_type': report_type,
        'page_title': 'Reports',
    }

    if report_type == 'monthly':
        context.update(_monthly_report(year, month))
    elif report_type == 'category':
        context.update(_category_report(year, month))
    elif report_type == 'account':
        context.update(_account_report(year, month))
    elif report_type == 'annual':
        context.update(_annual_report(year))

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(now.year - 3, now.year + 2))
    context['months'] = months
    context['years'] = years

    return render(request, 'reports/reports.html', context)


def _monthly_report(year, month):
    """Monthly breakdown report"""
    from transactions.services import get_monthly_summary, get_category_breakdown

    summary = get_monthly_summary(year, month)
    expense_cats = get_category_breakdown('expense', year, month)

    # Daily breakdown
    from datetime import date, timedelta
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    daily = Transaction.objects.filter(
        transaction_date__range=[first_day, last_day]
    ).values('transaction_date', 'transaction_type').annotate(
        total=Sum('amount')
    ).order_by('transaction_date')

    daily_map = {}
    for d in daily:
        key = str(d['transaction_date'])
        if key not in daily_map:
            daily_map[key] = {'income': 0, 'expense': 0}
        if d['transaction_type'] in ('income', 'expense'):
            daily_map[key][d['transaction_type']] = float(d['total'])

    current = first_day
    daily_labels, daily_income, daily_expense = [], [], []
    while current <= last_day:
        key = str(current)
        daily_labels.append(current.strftime('%d'))
        daily_income.append(daily_map.get(key, {}).get('income', 0))
        daily_expense.append(daily_map.get(key, {}).get('expense', 0))
        current += timedelta(days=1)

    chart_data = {
        'daily': {
            'labels': daily_labels,
            'income': daily_income,
            'expense': daily_expense,
        },
        'expense_pie': {
            'labels': [c['category__name'] or 'Other' for c in expense_cats],
            'data': [float(c['total']) for c in expense_cats],
            'colors': [c['category__color'] or '#6366f1' for c in expense_cats],
        },
    }

    monthly_summary = {
        'total_income': summary['income'],
        'total_expense': summary['expense'],
        'net_savings': summary['savings'],
        'savings_rate': summary['savings_rate'],
    }

    return {
        'monthly_summary': monthly_summary,
        'expense_cats': expense_cats,
        'chart_data_json': json.dumps(chart_data, default=decimal_default),
        'month_name': calendar.month_name[month],
    }


def _category_report(year, month):
    """Category-wise report (expenses)"""
    rows = Transaction.objects.filter(
        transaction_type='expense',
        transaction_date__year=year,
        transaction_date__month=month
    ).values(
        'category__name', 'category__color', 'category__icon'
    ).annotate(
        total=Sum('amount'), count=Count('id'), avg=Avg('amount')
    ).order_by('-total')

    grand_total = sum(float(r['total']) for r in rows) or 0

    category_data = []
    for r in rows:
        total = float(r['total'])
        category_data.append({
            'name': r['category__name'] or 'Uncategorized',
            'color': r['category__color'] or '#6366f1',
            'icon': r['category__icon'] or 'bi-three-dots',
            'transaction_count': r['count'],
            'total': total,
            'average': float(r['avg']) if r['avg'] is not None else 0,
            'percentage': (total / grand_total * 100) if grand_total else 0,
        })

    return {
        'category_data': category_data,
        'month_name': calendar.month_name[month],
    }


def _account_report(year, month):
    """Account-wise report"""
    accounts = BankAccount.objects.filter(is_active=True)
    account_data = []

    for acc in accounts:
        txn_count = Transaction.objects.filter(
            Q(bank_account=acc) | Q(target_account=acc),
            transaction_date__year=year,
            transaction_date__month=month
        ).count()

        account_data.append({
            'name': acc.name,
            'type': acc.get_account_type_display(),
            'color': acc.color,
            'balance': acc.balance,
            'transaction_count': txn_count,
        })

    return {
        'account_data': account_data,
        'month_name': calendar.month_name[month],
    }


def _annual_report(year):
    """Annual report with monthly breakdown"""
    monthly_data = []
    for m in range(1, 13):
        income = Transaction.objects.filter(
            transaction_type='income',
            transaction_date__year=year,
            transaction_date__month=m
        ).aggregate(t=Sum('amount'))['t'] or 0

        expense = Transaction.objects.filter(
            transaction_type='expense',
            transaction_date__year=year,
            transaction_date__month=m
        ).aggregate(t=Sum('amount'))['t'] or 0

        monthly_data.append({
            'month': calendar.month_abbr[m],
            'income': float(income),
            'expense': float(expense),
            'savings': float(income - expense),
        })

    chart_data = {
        'labels': [m['month'] for m in monthly_data],
        'income': [m['income'] for m in monthly_data],
        'expense': [m['expense'] for m in monthly_data],
        'savings': [m['savings'] for m in monthly_data],
    }

    annual_income = sum(m['income'] for m in monthly_data)
    annual_expense = sum(m['expense'] for m in monthly_data)

    return {
        'monthly_data': monthly_data,
        'annual_income': annual_income,
        'annual_expense': annual_expense,
        'annual_savings': annual_income - annual_expense,
        'chart_data_json': json.dumps(chart_data),
    }


def export_report(request):
    """Export filtered report"""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Report_{calendar.month_abbr[month]}_{year}'

    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    headers = ['Date', 'Type', 'Category', 'Account', 'Amount', 'Note']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    transactions = Transaction.objects.filter(
        transaction_date__year=year,
        transaction_date__month=month
    ).select_related('bank_account', 'category').order_by('transaction_date')

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=str(t.transaction_date))
        ws.cell(row=row, column=2, value=t.transaction_type)
        ws.cell(row=row, column=3, value=t.category.name if t.category else '')
        ws.cell(row=row, column=4, value=t.bank_account.name)
        ws.cell(row=row, column=5, value=float(t.amount))
        ws.cell(row=row, column=6, value=t.note)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'report_{calendar.month_abbr[month]}_{year}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
